import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MEDIUM = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FILL_HEADER = PatternFill("solid", fgColor="D9E2F3")
FILL_SUBHEADER = PatternFill("solid", fgColor="EDEDED")
FILL_COMPLETED = PatternFill("solid", fgColor="C6EFCE")
FILL_IN_PROGRESS = PatternFill("solid", fgColor="FFF2CC")
FILL_NOT_COMPLETED = PatternFill("solid", fgColor="FCE4D6")
FILL_BLOCKED = PatternFill("solid", fgColor="F4CCCC")
FILL_NOTE = PatternFill("solid", fgColor="D9D2E9")

FONT_BOLD = Font(bold=True)
FONT_NORMAL = Font(bold=False)
FONT_SMALL_BOLD = Font(bold=True, size=9)
FONT_SMALL = Font(bold=False, size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _safe_str(value):
    return "" if value is None else str(value).strip()


def _status_fill(status: str):
    status = _safe_str(status).lower()
    if status == "completed":
        return FILL_COMPLETED
    if status == "in_progress":
        return FILL_IN_PROGRESS
    if status == "blocked":
        return FILL_BLOCKED
    if status in {"not_completed", "failed"}:
        return FILL_NOT_COMPLETED
    return FILL_NOT_COMPLETED


def _prereq_to_text(prereq):
    if not prereq:
        return ""
    if isinstance(prereq, list):
        return ", ".join(str(x) for x in prereq if x)
    return str(prereq)


def _coreq_to_text(coreq):
    if not coreq:
        return ""
    if isinstance(coreq, list):
        return ", ".join(str(x) for x in coreq if x)
    return str(coreq)


def _semester_title(year_no: int, semester_no: int) -> str:
    year_map = {1: "First Year", 2: "Second Year", 3: "Third Year", 4: "Fourth Year"}
    sem_map = {1: "First Semester", 2: "Second Semester"}
    return f"{year_map.get(year_no, f'Year {year_no}')}: {sem_map.get(semester_no, f'Semester {semester_no}')}"


def _group_courses_by_semester(merged_rows):
    grouped = defaultdict(list)
    for row in merged_rows:
        year_no = row.get("year_no")
        semester_no = row.get("semester_no")
        if year_no and semester_no:
            grouped[(year_no, semester_no)].append(row)
    return grouped


def _write_semester_block(ws, start_row, start_col, title, courses):
    """
    Draw one semester block with 5 columns:
    code | title | credit | prerequisite | co requisite
    """
    end_col = start_col + 4

    # semester title
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    c = ws.cell(start_row, start_col, title)
    c.font = FONT_SMALL_BOLD
    c.alignment = CENTER
    c.fill = FILL_HEADER
    c.border = BORDER_MEDIUM

    # header row
    headers = ["Course code", "Course Title", "Credit", "Prerequisite", "Co Requisite"]
    for i, header in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + i, header)
        cell.font = FONT_SMALL_BOLD
        cell.alignment = CENTER
        cell.fill = FILL_SUBHEADER
        cell.border = BORDER_MEDIUM

    row_idx = start_row + 2
    total_credits = 0

    for course in courses:
        course_code = _safe_str(course.get("course_code"))
        course_name = _safe_str(course.get("course_name"))
        credit_hours = course.get("credit_hours") or ""
        prereq = _prereq_to_text(course.get("prerequisites"))
        coreq = _coreq_to_text(course.get("co_requisites"))
        status = _safe_str(course.get("status")).lower()
        grade = _safe_str(course.get("grade"))
        transcript_code = _safe_str(course.get("matched_transcript_code"))

        display_title = course_name
        if status == "completed" and grade:
            display_title = f"{course_name}\n[{grade}]"
        elif status == "in_progress":
            display_title = f"{course_name}\n[In Progress]"
        elif status == "blocked":
            display_title = f"{course_name}\n[Blocked]"
        elif transcript_code:
            display_title = f"{course_name}\n[{transcript_code}]"

        values = [course_code, display_title, credit_hours, prereq, coreq]
        fill = _status_fill(status)

        for i, value in enumerate(values):
            cell = ws.cell(row_idx, start_col + i, value)
            cell.font = FONT_SMALL
            cell.alignment = CENTER if i != 1 else LEFT
            cell.fill = fill
            cell.border = BORDER_THIN

        try:
            total_credits += int(credit_hours)
        except Exception:
            pass

        row_idx += 1

    # total row
    ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 1)
    total_label = ws.cell(row_idx, start_col, "Total")
    total_label.font = FONT_SMALL_BOLD
    total_label.alignment = CENTER
    total_label.border = BORDER_MEDIUM
    total_label.fill = FILL_SUBHEADER

    total_value = ws.cell(row_idx, start_col + 2, total_credits)
    total_value.font = FONT_SMALL_BOLD
    total_value.alignment = CENTER
    total_value.border = BORDER_MEDIUM
    total_value.fill = FILL_SUBHEADER

    for col in range(start_col + 3, end_col + 1):
        cell = ws.cell(row_idx, col, "")
        cell.border = BORDER_MEDIUM
        cell.fill = FILL_SUBHEADER

    # outer border reinforcement
    for r in range(start_row, row_idx):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = BORDER_THIN

    for c in range(start_col, end_col + 1):
        ws.cell(start_row, c).border = BORDER_MEDIUM
        ws.cell(start_row + 1, c).border = BORDER_MEDIUM
        ws.cell(row_idx, c).border = BORDER_MEDIUM

    for r in range(start_row, row_idx + 1):
        ws.cell(r, start_col).border = Border(left=MEDIUM, right=THIN, top=ws.cell(r, start_col).border.top, bottom=ws.cell(r, start_col).border.bottom)
        ws.cell(r, end_col).border = Border(left=THIN, right=MEDIUM, top=ws.cell(r, end_col).border.top, bottom=ws.cell(r, end_col).border.bottom)

    return row_idx


def build_structured_study_plan_workbook(
    merged_rows,
    output_path,
    program_name="Study Plan",
    total_required_credits=132,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Study Plan"

    # Column widths for two side-by-side semester blocks
    widths = {
        1: 12, 2: 28, 3: 10, 4: 14, 5: 14,
        6: 3,
        7: 12, 8: 28, 9: 10, 10: 14, 11: 14,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # top title
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"{program_name} Study Plan Audit"
    c.font = Font(bold=True, size=14)
    c.alignment = CENTER
    c.fill = FILL_HEADER
    c.border = BORDER_MEDIUM

    grouped = _group_courses_by_semester(merged_rows)

    current_row = 3

    semester_pairs = [
        ((1, 1), (1, 2)),
        ((2, 1), (2, 2)),
        ((3, 1), (3, 2)),
        ((4, 1), (4, 2)),
    ]

    for left_sem, right_sem in semester_pairs:
        left_courses = grouped.get(left_sem, [])
        right_courses = grouped.get(right_sem, [])

        left_title = _semester_title(*left_sem)
        right_title = _semester_title(*right_sem)

        left_height = len(left_courses) + 3
        right_height = len(right_courses) + 3
        block_height = max(left_height, right_height)

        _write_semester_block(ws, current_row, 1, left_title, left_courses)
        _write_semester_block(ws, current_row, 7, right_title, right_courses)

        # spacer row after pair
        current_row += block_height + 1

    # summer practical training row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    c = ws.cell(current_row, 1)
    c.value = "Summer Practical Training / AI 394 / Pre-requisite: Year 3 Core Courses / 1 Credit Hour"
    c.font = FONT_SMALL_BOLD
    c.alignment = CENTER
    c.fill = FILL_SUBHEADER
    c.border = BORDER_MEDIUM
    current_row += 1

    # total required credits
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    c1 = ws.cell(current_row, 1)
    c1.value = "Total Credits Required:"
    c1.font = FONT_SMALL_BOLD
    c1.alignment = CENTER
    c1.fill = FILL_SUBHEADER
    c1.border = BORDER_MEDIUM

    ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row, end_column=11)
    c2 = ws.cell(current_row, 9)
    c2.value = total_required_credits
    c2.font = Font(bold=True, color="FF0000")
    c2.alignment = CENTER
    c2.fill = FILL_SUBHEADER
    c2.border = BORDER_MEDIUM
    current_row += 2

    # note row
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + 1, end_column=11)
    note = ws.cell(current_row, 2)
    note.value = "Note: AI493 Co-op will replace the three corresponding courses in the coop plan (i.e. three elective courses in level 8)."
    note.font = FONT_SMALL_BOLD
    note.alignment = CENTER
    note.fill = FILL_NOTE
    note.border = BORDER_MEDIUM

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
    note_label = ws.cell(current_row, 1)
    note_label.value = "Note:"
    note_label.font = FONT_SMALL_BOLD
    note_label.alignment = CENTER
    note_label.fill = PatternFill("solid", fgColor="FF0000")
    note_label.border = BORDER_MEDIUM

    # row heights
    for r in range(1, current_row + 2):
        ws.row_dimensions[r].height = 28

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path
