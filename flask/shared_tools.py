import os

import requests


def _gemini_api_url(model: str, api_key: str) -> str:
    return f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"


def _extract_gemini_text(data: dict) -> str:
    candidates = data.get("candidates") or []
    for candidate in candidates:
        content = candidate.get("content") or {}
        for part in content.get("parts") or []:
            text = (part.get("text") or "").strip()
            if text:
                return text
    return ""


def generate_llm_response(prompt: str, fallback_text: str | None = None, timeout: int = 45) -> str:
    api_key = (os.environ.get("GEMINI_API_KEY") or "").strip()
    model = (os.environ.get("GEMINI_MODEL") or "gemini-2.5-flash").strip()

    if not api_key:
        return fallback_text or "Summary unavailable right now."

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.3,
            "topP": 0.9,
            "maxOutputTokens": 1024,
        },
    }

    try:
        res = requests.post(
            _gemini_api_url(model, api_key),
            json=payload,
            timeout=timeout,
        )
        if res.status_code != 200:
            return fallback_text or "Summary unavailable right now."

        try:
            data = res.json()
        except ValueError:
            return fallback_text or "Summary unavailable right now."

        text = _extract_gemini_text(data)
        return text or (fallback_text or "Summary unavailable right now.")
    except requests.exceptions.Timeout:
        return fallback_text or "Summary unavailable right now."
    except requests.exceptions.RequestException:
        return fallback_text or "Summary unavailable right now."
    except Exception:
        return fallback_text or "Summary unavailable right now."


def build_upload_summary_fallback(summary: dict) -> str:
    major = summary.get("major") or "Not detected"
    gpa = summary.get("gpa_final")
    progress = summary.get("progress_percent")
    completed = summary.get("completed_courses") or []
    in_progress = summary.get("in_progress_courses") or []
    remaining = summary.get("remaining_courses") or []
    next_courses = ", ".join(remaining[:5]) if remaining else "None"

    gpa_text = f"{float(gpa):.2f}" if isinstance(gpa, (int, float)) else (str(gpa) if gpa is not None else "N/A")
    progress_text = f"{float(progress):.2f}%" if isinstance(progress, (int, float)) else (str(progress) if progress is not None else "N/A")

    return (
        f"- Major: {major}\n"
        f"- Undergraduate GPA: {gpa_text}\n"
        f"- Progress: {progress_text}\n"
        f"- Courses completed: {len(completed)}; in progress: {len(in_progress)}\n"
        f"- Next remaining courses: {next_courses}"
    )


def build_chat_fallback(facts: dict, message: str) -> str:
    msg = (message or "").lower().strip()
    completed = facts.get("completed") or []
    in_progress = facts.get("in_progress") or []
    remaining = facts.get("remaining") or []
    progress = facts.get("progress")
    gpa_prep = facts.get("gpa_prep") or "N/A"
    gpa_ug = facts.get("gpa_ug") or "N/A"
    major = facts.get("major") or "Not detected"

    if "gpa" in msg:
        return f"Preparatory GPA: {gpa_prep}\nUndergraduate GPA: {gpa_ug}"
    if "remaining" in msg or "left" in msg:
        return f"Remaining courses ({len(remaining)}): {', '.join(remaining) if remaining else 'None'}"
    if "progress" in msg or "%" in msg:
        return f"Your overall program progress is {progress}%."
    if "register" in msg or "next" in msg:
        return f"Recommended next courses: {', '.join(remaining[:3]) if remaining else 'All courses completed.'}"
    if "completed" in msg or "finished" in msg or "how many" in msg:
        total = len(completed) + len(in_progress) + len(remaining)
        return f"You have completed {len(completed)} courses out of {total} total."

    return (
        f"Major: {major}\n"
        f"Progress: {progress}%\n"
        f"GPA (Prep/UG): {gpa_prep}/{gpa_ug}\n"
        f"Completed: {len(completed)}, In progress: {len(in_progress)}, Remaining: {len(remaining)}"
    )


def compare_transcript_with_plan(major_name, df_transcript, study_plan_data, extracted_summary_df=None):
    """
    FINAL PRODUCTION VERSION (2025-10-30)
    Fully JSON-driven.
    """

    import pandas as pd
    from collections import Counter

    def norm(x: str) -> str:
        return "".join(str(x or "").upper().split())

    def is_slot(x: str) -> bool:
        return norm(x).endswith("XXX")

    programs = study_plan_data.get("programs", study_plan_data)
    program = programs.get(major_name)
    if not program:
        raise ValueError(f"No study plan found for {major_name}")

    print(f"\nComparing transcript against study plan for: {major_name}")

    plan_levels = program["levels"]
    slot_rules = program.get("slot_rules", {})
    plan_codes = [norm(c["course_code"]) for lvl in plan_levels.values() for c in lvl]

    df_transcript = df_transcript.copy()

    def classify_status(grade, points):
        g = str(grade or "").upper().strip()
        if g in {"IP", "INPROGRESS", "IN PROGRESS"}:
            return "in_progress"
        if g in {"F", "FA", "NF"}:
            return "not_completed"
        try:
            if float(points) > 0:
                return "completed"
        except Exception:
            pass
        return "in_progress"

    df_transcript["Status"] = df_transcript.apply(
        lambda r: classify_status(r.get("Grade"), r.get("Points")), axis=1
    )

    completed = set(df_transcript[df_transcript["Status"] == "completed"]["Course Code"].map(norm))
    in_progress = set(df_transcript[df_transcript["Status"] == "in_progress"]["Course Code"].map(norm))
    failed = set(df_transcript[df_transcript["Status"] == "not_completed"]["Course Code"].map(norm))

    prep = {"ENGL000","ENGL001","ENGL002","ENGL003","ENGL004","ENGL005","MATH001","MATH002","PCS001","PCD001"}
    completed -= prep
    in_progress -= prep
    failed -= prep

    removed_from_plan = set()
    if "AI493" in completed or "AI493" in in_progress:
        removed_from_plan.update(["AIXXX", "C3SXXX"])
        effective_plan = [c for c in plan_codes if c not in {"AIXXX","C3SXXX"}]
    else:
        removed_from_plan.add("AI493")
        effective_plan = [c for c in plan_codes if c != "AI493"]

    effective_plan_set = set(effective_plan)
    plan_core = {c for c in effective_plan if not is_slot(c)}
    slot_need = Counter([c for c in effective_plan if is_slot(c)])

    equiv_sets = {
        "GHALXXX": {"GHALXXX","GIASXXX","GIAS101","GIAS102","GIAS221"},
        "GIASXXX": {"GIASXXX","GHALXXX","GIAS101","GIAS102","GIAS221"},
    }

    def canonical_slot_for(code: str) -> str:
        u = norm(code)
        if u in effective_plan_set and is_slot(u):
            return u
        if u in equiv_sets["GHALXXX"] or u in equiv_sets["GIASXXX"]:
            if "GHALXXX" in effective_plan_set:
                return "GHALXXX"
            if "GIASXXX" in effective_plan_set:
                return "GIASXXX"
        return u

    def infer_slot(course_code: str):
        code = norm(course_code)
        if code in plan_core:
            return None

        possible_slots = []
        for slot, rule in slot_rules.items():
            s = norm(slot)
            if s not in effective_plan_set:
                continue
            specifics = {norm(x) for x in rule.get("specific", []) if x}
            prefixes = {norm(x) for x in rule.get("prefixes", []) if x}

            if code in specifics:
                possible_slots.append((s, "specific"))
            elif any(code.startswith(p) for p in prefixes):
                possible_slots.append((s, "prefix"))

        if not possible_slots:
            if code in equiv_sets["GHALXXX"] or code in equiv_sets["GIASXXX"]:
                return canonical_slot_for(code)
            return None

        specific_matches = [s for s, t in possible_slots if t == "specific"]
        if specific_matches:
            return specific_matches[0]
        if len(possible_slots) > 1:
            ranked = sorted(possible_slots, key=lambda kv: len(slot_rules[kv[0]].get("prefixes", [])))
            return ranked[0][0]
        return possible_slots[0][0]

    slot_mapping = {"completed": {}, "in_progress": {}}
    slot_done = Counter()
    slot_doing = Counter()

    for c in completed:
        s = infer_slot(c)
        if s:
            s = canonical_slot_for(s)
            slot_mapping["completed"][c] = s
            slot_done[s] += 1

    for c in in_progress:
        s = infer_slot(c)
        if s:
            s = canonical_slot_for(s)
            slot_mapping["in_progress"][c] = s
            slot_doing[s] += 1

    remaining = [c for c in plan_core if c not in completed and c not in in_progress]
    for slot_code, need in slot_need.items():
        filled = slot_done[slot_code] + slot_doing[slot_code]
        if filled < need:
            remaining.extend([slot_code] * (need - filled))
    remaining = sorted(set(remaining))

    def metric_val(name):
        if not isinstance(extracted_summary_df, pd.DataFrame) or extracted_summary_df.empty:
            return None

        cols = [c.lower() for c in extracted_summary_df.columns]
        if "metric" not in cols and "cumulative gpa" in cols:
            try:
                return extracted_summary_df["Cumulative GPA"].astype(float).max()
            except Exception:
                return None

        if "metric" in cols:
            mask = extracted_summary_df["Metric"].astype(str).str.lower() == name.lower()
            if mask.any():
                try:
                    return extracted_summary_df.loc[mask, "Value"].astype(float).iloc[0]
                except Exception:
                    pass
        return None

    completed_view = [
        f"{c} (counts as {slot_mapping['completed'][c]})" if c in slot_mapping["completed"] else c
        for c in sorted(completed)
    ]
    inprogress_view = [
        f"{c} (counts as {slot_mapping['in_progress'][c]})" if c in slot_mapping["in_progress"] else c
        for c in sorted(in_progress)
    ]

    cores_done = len([c for c in plan_core if c in completed])
    slots_filled = sum(min(slot_done[s] + slot_doing[s], slot_need[s]) for s in slot_need)
    total_required = len(plan_core) + sum(slot_need.values())
    progress_percent = round(((cores_done + slots_filled) / total_required) * 100, 2) if total_required else 0.0

    return {
        "major": major_name,
        "completed_courses": completed_view,
        "in_progress_courses": inprogress_view,
        "remaining_courses": remaining,
        "removed_from_plan": sorted(removed_from_plan),
        "gpa_prep": metric_val("Preparatory GPA"),
        "gpa_final": metric_val("Final Cumulative GPA"),
        "credit_prep": metric_val("Preparatory Credits"),
        "credit_ug": metric_val("Undergraduate Credits"),
        "progress_percent": progress_percent,
    }


def generate_structured_study_plan_excel(student_info, summary, study_plan_data, output_path="uploads/Student_Summary_Report.xlsx"):
    import os
    import re
    import openpyxl
    from collections import defaultdict, Counter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    equivalent_map = {
        "AIXXX": [],
        "C3SXXX": ["CS314"],
        "GHALXXX": ["GIAS221", "GIASXXX"],
        "GIASXXX": ["GHALXXX", "GIAS221"],
        "GSOSXXX": ["GSOS214"],
    }

    def canonical_slot(code):
        code = (code or "").upper().replace(" ", "")
        for slot, eqs in equivalent_map.items():
            if code == slot or code in eqs:
                return slot
        return code

    def norm(x): return str(x or "").upper().replace(" ", "")
    def is_slot(x): return norm(x).endswith("XXX")

    strip_paren = lambda s: norm(s.split("(")[0])
    completed_raw = summary.get("completed_courses", [])
    inprog_raw = summary.get("in_progress_courses", [])
    failed_raw = summary.get("failed_courses", [])
    remaining_raw = summary.get("remaining_courses", [])
    removed_from_plan = set(map(norm, summary.get("removed_from_plan", [])))

    completed = {strip_paren(c) for c in completed_raw}
    inprogress = {strip_paren(c) for c in inprog_raw}
    failed = {strip_paren(c) for c in failed_raw}
    remaining = {strip_paren(c) for c in remaining_raw}

    slot_mapping_completed = defaultdict(list)
    slot_mapping_inprog = defaultdict(list)
    pat = re.compile(r"\(.*?counts\s*as\s*([A-Za-z0-9]+)\)?", re.IGNORECASE)

    def parse_mapping(course_list, target):
        for c in course_list:
            m = pat.search(c)
            if not m:
                continue
            slot = canonical_slot(norm(m.group(1)))
            real = strip_paren(c)
            if canonical_slot(real) == canonical_slot(slot):
                continue
            target[slot].append(real)

    parse_mapping(completed_raw, slot_mapping_completed)
    parse_mapping(inprog_raw, slot_mapping_inprog)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Student_Summary"
    ws2 = wb.create_sheet("Study_Plan_Structured")

    colors = {
        "Completed": "C6EFCE",
        "In Progress": "FFF3CD",
        "Remaining": "F8D7DA",
        "Failed": "F5B7B1",
    }
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 70
    ws1.append(["Field", "Value"])
    ws1["A1"].font = ws1["B1"].font = Font(bold=True)

    gpa_prep = summary.get("gpa", {}).get("prep", "—")
    gpa_undergrad = summary.get("gpa", {}).get("undergrad", "—")
    totals = summary.get("totals", {})

    info_pairs = [
        ("Student Name", student_info.get("Name", "Unknown")),
        ("Major", summary.get("major", "BS in Artificial Intelligence and Data Science")),
        ("Preparatory GPA", gpa_prep),
        ("Undergraduate GPA", gpa_undergrad),
        ("Progress (%)", f"{summary.get('progress_percent', 0)}%"),
        ("Total Required Courses", totals.get("required", "—")),
        ("Completed Courses", totals.get("completed", len(completed))),
        ("In Progress", totals.get("in_progress", len(inprogress))),
        ("Remaining Courses", totals.get("remaining", len(remaining))),
        ("Academic Path", summary.get("academic_path", "Normal Elective")),
    ]

    for a, b in info_pairs:
        ws1.append([a, b])
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    used_courses = set()

    slot_occurrences = Counter()
    for year, semesters in study_plan_data.items():
        for sem, courses in semesters.items():
            for c in courses:
                code = canonical_slot(norm(c.get("course_code", "")))
                if is_slot(code):
                    slot_occurrences[code] += 1

    def slot_status(slot_code):
        cslot = canonical_slot(slot_code)

        if not is_slot(cslot):
            if cslot in completed:
                return "Completed", ""
            if cslot in inprogress:
                return "In Progress", ""
            if cslot in failed:
                return "Failed", ""
            return "Remaining", ""

        mapping_pool = slot_mapping_completed.get(cslot, []) or slot_mapping_inprog.get(cslot, [])
        if mapping_pool:
            real = mapping_pool.pop(0)
            used_courses.add(real)
            if real in completed:
                return "Completed", f" -> Satisfied by {real}"
            if real in inprogress:
                return "In Progress", f" -> In Progress ({real})"

        for pool, status in [(completed, "Completed"), (inprogress, "In Progress")]:
            for eq in equivalent_map.get(cslot, []):
                if eq in pool and eq not in used_courses:
                    used_courses.add(eq)
                    return status, f" -> Satisfied by {eq}"

        return "Remaining", ""

    row_idx = 1
    for year, semesters in study_plan_data.items():
        for sem, courses in semesters.items():
            ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            h = ws2.cell(row=row_idx, column=1, value=f"{year}: {sem}")
            h.font = Font(bold=True, size=14)
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.fill = PatternFill("solid", fgColor="D9EAD3")
            row_idx += 1

            headers = ["Course Code", "Course Title", "Credit Hours", "Prerequisite", "Co-Requisite", "Status"]
            for i, col in enumerate(headers, 1):
                c = ws2.cell(row=row_idx, column=i, value=col)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")
                c.fill = PatternFill("solid", fgColor="E2EFDA")
                c.border = border
            row_idx += 1

            total_credits = 0.0
            for c in courses:
                code = canonical_slot(norm(c.get("course_code", "")))
                if code in removed_from_plan:
                    continue
                title = c.get("title", "")
                credit = c.get("credits", "")
                pre = c.get("prerequisite", "") or "-"
                co = c.get("co_requisite", "") or "-"

                status, note = slot_status(code)
                if note:
                    title += note
                try:
                    total_credits += float(credit or 0)
                except Exception:
                    pass

                row_data = [code, title, credit, pre, co, status]
                for i, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=row_idx, column=i, value=val)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                    if i == 6:
                        cell.fill = PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))
                row_idx += 1

            ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            ws2.cell(row=row_idx, column=1, value="Total Credits:").font = Font(bold=True)
            ws2.cell(row=row_idx, column=6, value=total_credits).font = Font(bold=True)
            row_idx += 2

    legend = ["Legend:", "Completed", "In Progress", "Remaining", "Failed"]
    ws2.append(legend)
    legend_row = ws2.max_row
    for i, label in enumerate(legend[1:], 2):
        c = ws2.cell(row=legend_row, column=i, value=label)
        c.fill = PatternFill("solid", fgColor=colors[label])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for col in range(1, 8):
        ws2.column_dimensions[chr(64 + col)].width = 18

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
