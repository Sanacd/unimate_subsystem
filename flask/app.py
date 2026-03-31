import sqlite3
import os, sys
import json
import re
import requests
import traceback
from flask import Flask, request, jsonify, send_from_directory
import webbrowser
from flask import send_file
from flask import send_file, after_this_request
import sqlite3, json, os
from sentence_transformers import SentenceTransformer

def get_db_path():
    # Running as EXE (PyInstaller)
    if hasattr(sys, "_MEIPASS"):
        # Windows AppData folder for UniMate
        appdata = os.getenv("LOCALAPPDATA") or os.path.expanduser("~\\AppData\\Local")
        app_dir = os.path.join(appdata, "UniMate")
        os.makedirs(app_dir, exist_ok=True)
        return os.path.join(app_dir, "agent_runtime.db")

    # ✅ Normal dev mode → keep DB inside backend folder
    return os.path.join(os.getcwd(), "agent_runtime.db")


DB_FILE = get_db_path()

# ------------------------------------------------------
# ✅ Initialize DB if not exists
# ------------------------------------------------------
def init_agent_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS sessions (
        id TEXT PRIMARY KEY,
        created_at REAL,
        user_id TEXT,
        reasoning_mode TEXT,
        hitl_enabled INTEGER,
        orchestrator_mode TEXT
    )""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS agent_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id TEXT,
        agent TEXT,
        role TEXT,
        content TEXT,
        ts REAL
    )""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS hitl_queue (
        id TEXT PRIMARY KEY,
        session_id TEXT,
        agent TEXT,
        action TEXT,
        payload_json TEXT,
        status TEXT DEFAULT 'pending',
        created_at REAL
    )""")

    # ✅ Unified state table (for session memory)
    c.execute("""
    CREATE TABLE IF NOT EXISTS state (
        id TEXT PRIMARY KEY,
        session_id TEXT,
        key TEXT,
        value_json TEXT
    )""")

    # ✅ Artifacts table (Excel reports etc.)
    c.execute("""
    CREATE TABLE IF NOT EXISTS artifacts (
        id TEXT PRIMARY KEY,
        session_id TEXT,
        kind TEXT,
        path TEXT,
        meta_json TEXT
    )""")

    conn.commit()
    conn.close()


# Initialize DB immediately
init_agent_db()

print(f"✅ DB Path: {DB_FILE}")

# --------------------------------------------------------------------
# ✅ Initialize Flask App & Path Resolution
# --------------------------------------------------------------------
app = Flask(__name__, static_folder=None)  # we control static manually
# ------------------------------------------------------
# 🔹 Embedding API for RAG (/api/embed)
# ------------------------------------------------------
EMB_MODEL = "sentence-transformers/distiluse-base-multilingual-cased-v1"
embedder = SentenceTransformer(EMB_MODEL)

@app.post("/api/embed")
# ========== Embedding API ==========
@app.post("/embed")
def embed_api():
    try:
        data = request.get_json(force=True) or {}
        text = (data.get("text") or "").strip()

        if not text:
            return jsonify({"error": "empty text"}), 400

        # SentenceTransformer expects a LIST
        vector = embedder.encode([text])[0]

        return jsonify({"embedding": vector.tolist()})

    except Exception as e:
        print("EMB_ERR:", e)
        return jsonify({"error": str(e)}), 500


# Detect bundled vs dev
IS_BUNDLED = hasattr(sys, "_MEIPASS")
BASE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))

# ✅ Frontend & static folders (inside backend)
FRONTEND_DIR = os.path.join(BASE_DIR, "..", "node", "public")
STATIC_DIR   = os.path.join(BASE_DIR, "..", "node", "public")

# ✅ Uploads folder logic
if IS_BUNDLED:
    # EXE → user home AppData folder for safety & write permissions
    UPLOAD_FOLDER = os.path.join(os.getenv("LOCALAPPDATA"), "UniMate", "uploads")
else:
    # Dev mode → backend/uploads
    UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --------------------------------------------------------------------
# Logging helper
# --------------------------------------------------------------------
def debug_log(msg):
    with open("debug_log.txt", "a", encoding="utf-8") as f:
        f.write(msg + "\n")

debug_log("🚀 App started")
debug_log(f"MODE: {'EXE' if IS_BUNDLED else 'DEV'}")
debug_log(f"BASE_DIR: {BASE_DIR}")
debug_log(f"FRONTEND_DIR: {FRONTEND_DIR}")
debug_log(f"STATIC_DIR: {STATIC_DIR}")
debug_log(f"UPLOAD_FOLDER: {UPLOAD_FOLDER}")

# --------------------------------------------------------------------
# ✅ Routes to serve UI
# --------------------------------------------------------------------
@app.route("/")
def serve_index():
    index_file = os.path.join(FRONTEND_DIR, "index.html")
    if os.path.exists(index_file):
        return send_from_directory(FRONTEND_DIR, "index.html")
    return "<h3>❌ index.html not found in /frontend</h3>", 404

@app.route("/static/<path:filename>")
def serve_static_folder(filename):
    file_path = os.path.join(STATIC_DIR, filename)
    if os.path.exists(file_path):
        return send_from_directory(STATIC_DIR, filename)
    return "<h3>Static file not found</h3>", 404


@app.route("/<path:asset>")
def serve_files(asset):
    # Skip /static paths — they are handled above
    if asset.startswith("static/"):
        return serve_static_folder(asset.replace("static/", ""))

    frontend_path = os.path.join(FRONTEND_DIR, asset)
    static_path = os.path.join(STATIC_DIR, asset)

    if os.path.exists(frontend_path):
        return send_from_directory(FRONTEND_DIR, asset)
    if os.path.exists(static_path):
        return send_from_directory(STATIC_DIR, asset)

    return "<h3>❌ File not found</h3>", 404


@app.route("/ask", methods=["POST"])
def handle_ask():
    data = request.get_json(force=True) or {}
    message = data.get("message", "")
    role = data.get("role", "guest")

    
    return jsonify({"reply": "OK"})

# --------------------------------------------------------------------
# 📚 Load Study Plan JSON (safe + flexible)
# --------------------------------------------------------------------
studyplan_path = os.path.join(BASE_DIR, "university_studyplans.json")

policy_path = os.path.join(BASE_DIR, "academic_policy.json")
with open(policy_path, "r", encoding="utf-8") as f:
    ACADEMIC_POLICY = json.load(f)


if not os.path.exists(studyplan_path):
    raise FileNotFoundError(
        f"❌ Study plan JSON not found at {studyplan_path}\n"
        f"Make sure university_studyplans.json is inside backend/"
    )

with open(studyplan_path, "r", encoding="utf-8") as f:
    try:
        STUDY_PLANS = json.load(f).get("programs", {})
        if not STUDY_PLANS:
            raise ValueError("JSON loaded but no `programs` key found.")
    except Exception as e:
        raise RuntimeError(f"❌ Failed to parse study plan JSON: {e}")

student_context = {}
print(f"✅ Study plans loaded — {len(STUDY_PLANS)} programs found")


# --------------------------------------------------------------------
# ⚙️ Helpers
# --------------------------------------------------------------------
def norm(code: str) -> str:
    """Normalize course codes (e.g., ' cs 314 ' → 'CS314')"""
    return "".join(str(code or "").upper().split())


def classify_status(grade_raw, points_raw) -> str:
    """
    Determine course completion state from grade + points.
    Neutral, consistent classification.
    """
    g = str(grade_raw or "").strip().upper()

    try:
        pts = float(points_raw or 0)
    except ValueError:
        pts = 0

    if g in {"W", "WITHDRAWN"}:       return "withdrawn"
    if g in {"WAIVED", "WAIVE"}:      return "completed"
    if g in {"F", "FA", "NF"}:        return "not_completed"
    if pts == 0 and g in {"", "0", "0.0", "0.00"}:
        return "in_progress"
    return "completed"


# --------------------------------------------------------------------
# 🔁 Equivalence & Normalization Enhancements
# --------------------------------------------------------------------
# Maintain canonical dictionary for slot equivalence
EQUIVALENT_MAP = {
    "GIASXXX": ["GHALXXX", "GIAS221"],
    "GHALXXX": ["GIASXXX", "GIAS221"],
    "C3SXXX":  ["CS314"],      # College elective
    "GSOSXXX": ["GSOS214"],   # Social science elective
    "AIXXX":   [],            # Program elective slot — may map later
}

def expand_equivalents(courses: set, mapping: dict) -> set:
    """
    Expand course slots based on equivalence map.
    Example input: {"GIAS221"} → {"GIAS221","GIASXXX","GHALXXX"}
    """
    expanded = set(courses)

    for code in list(courses):
        for slot, equivalents in mapping.items():
            if code == slot or code in equivalents:
                expanded.add(slot)
                expanded.update(equivalents)

    return expanded

def recognize_major_by_overlap(df_transcript, study_plans, debug=False):
    """
    Detect student's major by finding max overlap between transcript course codes
    and each program's official study plan.

    Supports:
    ✔ fallback logic
    ✔ debug trace
    ✔ defensive handling of bad / empty DataFrames
    """

    # Validate df
    if df_transcript is None or "Course Code" not in df_transcript.columns:
        if debug: print("⚠️ No transcript data or missing 'Course Code' column.")
        return "Unknown Major"

    student_codes = {norm(c) for c in df_transcript["Course Code"].dropna().tolist()}
    if debug: print(f"🎓 Student course codes detected: {len(student_codes)} items")

    best_major, best_hit = None, -1
    tie_candidates = []

    for major_name, plan in study_plans.items():
        plan_levels = plan.get("levels", {})
        plan_codes = {
            norm(c.get("course_code", ""))
            for lvl in plan_levels.values()
            for c in lvl
        }

        overlap = len(student_codes & plan_codes)

        if debug:
            print(f"📚 Checking major '{major_name}': {overlap} overlap matches")

        if overlap > best_hit:
            best_hit = overlap
            best_major = major_name
            tie_candidates = [(major_name, overlap)]
        elif overlap == best_hit:
            tie_candidates.append((major_name, overlap))

    # ✅ Tie-breaking logic
    if len(tie_candidates) > 1:
        if debug:
            print(f"⚖️ Tie detected: {tie_candidates}")

        # heuristic: prefer majors that include AI courses
        priority_keywords = ["artificial", "cyber", "computer", "software", "ai"]
        for major, _ in tie_candidates:
            if any(k in major.lower() for k in priority_keywords):
                if debug: print(f"🏆 Tie broken → {major}")
                best_major = major
                break

    # ✅ Final fallback
    if not best_major or best_hit == 0:
        if debug:
            print("⚠️ No major overlap, defaulting → BS in Artificial Intelligence and Data Science")
        return "BS in Artificial Intelligence and Data Science"

    if debug:
        print(f"✅ Final detected major: {best_major} (overlap = {best_hit})")

    return best_major

def compare_transcript_with_plan(major_name, df_transcript, study_plan_data, extracted_summary_df=None):
    """
    FINAL STABLE VERSION (Unified helpers)
    • Slot inference (AIXXX, C3SXXX, GHALXXX / GIASXXX)
    • Handles co-op vs elective paths
    • Uses global helpers + equivalence sets
    • Compatible with Excel generator
    """

    import pandas as pd
    from collections import Counter

    programs = study_plan_data.get("programs", study_plan_data)
    program = programs.get(major_name)

    if not program:
        raise ValueError(f"No study plan found for {major_name}")

    print(f"\n🧩 Comparing transcript against study plan for: {major_name}")

    plan_levels = program["levels"]
    slot_rules  = program.get("slot_rules", {})

    # ---------- plan flatten ----------
    plan_codes = [norm(c["course_code"]) for lvl in plan_levels.values() for c in lvl]
    plan_set   = set(plan_codes)

    # ---------- classify transcript ----------
    df_transcript = df_transcript.copy()

    df_transcript["Status"] = df_transcript.apply(
        lambda r: classify_status(r.get("Grade"), r.get("Points")),
        axis=1
    )

    completed     = set(df_transcript[df_transcript["Status"] == "completed"]["Course Code"].map(norm))
    in_progress   = set(df_transcript[df_transcript["Status"] == "in_progress"]["Course Code"].map(norm))
    failed        = set(df_transcript[df_transcript["Status"] == "not_completed"]["Course Code"].map(norm))

    # remove prep courses
    PREP = {"ENGL000","ENGL001","ENGL002","ENGL003","ENGL004","ENGL005",
            "MATH001","MATH002","PCS001","PCD001"}
    completed   -= PREP
    in_progress -= PREP
    failed      -= PREP

    # ---------- co-op path ----------
    removed_from_plan = set()
    if "AI493" in completed or "AI493" in in_progress:
        removed_from_plan.update(["AIXXX", "C3SXXX"])
        effective_plan = [c for c in plan_codes if c not in {"AIXXX", "C3SXXX"}]
        print("🔄 Co-op path — electives removed.")
    else:
        removed_from_plan.add("AI493")
        effective_plan = [c for c in plan_codes if c != "AI493"]
        print("🟡 Elective path — co-op removed.")

    effective_plan_set = set(effective_plan)
    plan_core = {c for c in effective_plan if not c.endswith("XXX")}

    slot_need = Counter([c for c in effective_plan if c.endswith("XXX")])

    # ---------- canonical slot resolver ----------
    def canonical_slot_for(code: str):
        u = norm(code)
        eq = expand_equivalents({u}, EQUIVALENT_MAP)
        valid = eq & effective_plan_set
        return next(iter(valid)) if valid else u

    # ---------- slot inference (JSON-driven) ----------
    def infer_slot(course_code: str):
        code = norm(course_code)
        if code in plan_core:
            return None

        # collect candidates
        possible = []
        for slot, rule in slot_rules.items():
            s = norm(slot)
            if s not in effective_plan_set:
                continue

            specifics = {norm(x) for x in rule.get("specific", []) if x}
            prefixes  = {norm(x) for x in rule.get("prefixes", []) if x}

            if code in specifics:
                possible.append((s, "specific"))
            elif any(code.startswith(p) for p in prefixes):
                possible.append((s, "prefix"))

        if not possible:
            return canonical_slot_for(code) if code in expand_equivalents({code}, EQUIVALENT_MAP) else None

        # priority: specific > most-focused prefix
        specific = [s for s, t in possible if t == "specific"]
        if specific:
            return specific[0]

        # tie-break: fewer prefixes = more specific slot
        if len(possible) > 1:
            return sorted(possible, key=lambda kv: len(slot_rules[kv[0]].get("prefixes", [])))[0][0]

        return possible[0][0]

    # ---------- slot matching ----------
    slot_mapping = {"completed": {}, "in_progress": {}}
    slot_done, slot_doing = Counter(), Counter()

    for c in completed:
        slot = infer_slot(c)
        if slot:
            slot = canonical_slot_for(slot)
            slot_mapping["completed"][c] = slot
            slot_done[slot] += 1
            print(f"✅ [slot match] {c} → {slot}")

    for c in in_progress:
        slot = infer_slot(c)
        if slot:
            slot = canonical_slot_for(slot)
            slot_mapping["in_progress"][c] = slot
            slot_doing[slot] += 1
            print(f"🕓 [slot match] {c} → {slot}")

    # ---------- remaining ----------
    remaining = [c for c in plan_core if c not in completed and c not in in_progress]
    for slot_code, need in slot_need.items():
        have = slot_done[slot_code] + slot_doing[slot_code]
        if have < need:
            remaining.extend([slot_code] * (need - have))
    remaining = sorted(set(remaining))

    # ---------- GPA ----------
    def metric_val(name):
        if isinstance(extracted_summary_df, pd.DataFrame) and not extracted_summary_df.empty:
            d = extracted_summary_df[extracted_summary_df["Metric"].str.lower() == name.lower()]
            if not d.empty:
                try: return float(d["Value"].iloc[0])
                except: pass
        return 0.0

    # ---------- Output ----------
    summary = {
        "major": major_name,
        "completed_courses": [
            f"{c} (counts as {slot_mapping['completed'][c]})" if c in slot_mapping["completed"] else c
            for c in sorted(completed)
        ],
        "in_progress_courses": [
            f"{c} (counts as {slot_mapping['in_progress'][c]})" if c in slot_mapping["in_progress"] else c
            for c in sorted(in_progress)
        ],
        "remaining_courses": remaining,
        "removed_from_plan": sorted(removed_from_plan),
        "gpa_prep": metric_val("Preparatory GPA"),
        "gpa_final": metric_val("Final Cumulative GPA"),
        "credit_prep": metric_val("Preparatory Credits"),
        "credit_ug": metric_val("Undergraduate Credits"),
        "progress_percent": round(
            (
                (len([c for c in plan_core if c in completed]) +
                 sum(min(slot_done[s] + slot_doing[s], slot_need[s]) for s in slot_need))
                / (len(plan_core) + sum(slot_need.values()))
            ) * 100,
            2,
        ),
    }

    # ---------- Debug print ----------
    print("\n🔎 DEBUG SUMMARY")
    print(f"✅ Completed ({len(completed)}): {sorted(completed)}")
    print(f"🕓 In Progress ({len(in_progress)}): {sorted(in_progress)}")
    print(f"📘 Remaining ({len(remaining)}): {remaining}")
    if slot_need:
        print("📊 Slot Fill Summary:")
        for s, need in slot_need.items():
            print(f"   • {s}: {slot_done[s] + slot_doing[s]}/{need} filled "
                  f"(✅ {slot_done[s]}, 🕓 {slot_doing[s]})")
    print("=" * 80)

    return summary

def generate_structured_study_plan_excel(student_info, summary, study_plan_data, output_path="uploads/Student_Summary_Report.xlsx"):
    """
    FINAL STABLE VERSION — aligned with agent summary dictionary
    ✅ Reads GPA & totals from summary["gpa"] and summary["totals"]
    ✅ Uses global EQUIVALENT_MAP (no duplication)
    ✅ Consistent with compare_transcript_with_plan output
    """

    import os, re, openpyxl
    from collections import defaultdict, Counter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def canonical_slot(code): 
        code = (code or "").upper().replace(" ", "")
        for slot, eqs in EQUIVALENT_MAP.items():
            if code == slot or code in eqs:
                return slot
        return code

    def norm(x): return str(x or "").upper().replace(" ", "")
    def is_slot(x): return norm(x).endswith("XXX")
    strip_paren = lambda s: norm(s.split("(")[0])

    # ✅ Extract sets from summary
    completed_raw   = summary.get("completed_courses", [])
    inprog_raw      = summary.get("in_progress_courses", [])
    failed_raw      = summary.get("failed_courses", [])
    remaining_raw   = summary.get("remaining_courses", [])
    removed_plan    = set(map(norm, summary.get("removed_from_plan", [])))

    completed   = {strip_paren(c) for c in completed_raw}
    inprogress  = {strip_paren(c) for c in inprog_raw}
    failed      = {strip_paren(c) for c in failed_raw}
    remaining   = {strip_paren(c) for c in remaining_raw}

    # ✅ Detect mapped electives in text "(counts as X)"
    slot_mapping_completed = defaultdict(list)
    slot_mapping_inprog    = defaultdict(list)
    pat = re.compile(r"counts\s*as\s*([A-Za-z0-9]+)", re.IGNORECASE)

    def parse_mapping(course_list, bucket):
        for c in course_list:
            m = pat.search(c)
            if not m: continue
            slot = canonical_slot(norm(m.group(1)))
            real = strip_paren(c)
            if canonical_slot(real) != slot:
                bucket[slot].append(real)
                print(f"🧩 Mapping found: {real} → {slot}")

    parse_mapping(completed_raw, slot_mapping_completed)
    parse_mapping(inprog_raw, slot_mapping_inprog)

    # ✅ Workbook setup
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Student_Summary"
    ws2 = wb.create_sheet("Study_Plan_Structured")

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 70
    ws1.append(["Field", "Value"])
    ws1["A1"].font = ws1["B1"].font = Font(bold=True)

    # ✅ Correct GPA & totals
    gpa   = summary.get("gpa", {})
    totals = summary.get("totals", {})

    info_pairs = [
        ("Student Name", student_info.get("Name", "Unknown")),
        ("Major", summary.get("major", "AI Program")),
        ("Preparatory GPA", gpa.get("prep", "—")),
        ("Undergraduate GPA", gpa.get("undergrad", "—")),
        ("Progress (%)", f"{summary.get('progress_percent', 0)}%"),
        ("Total Required Courses", totals.get("required", "—")),
        ("Completed Courses", totals.get("completed", len(completed))),
        ("In Progress", totals.get("in_progress", len(inprogress))),
        ("Remaining Courses", totals.get("remaining", len(remaining))),
        ("Academic Path", summary.get("academic_path", "Normal Elective")),
    ]
    for a, b in info_pairs:
        ws1.append([a, b])
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    border = Border(*(Side(style="thin"),)*4)
    colors = {"✅ Completed": "C6EFCE", "🕓 In Progress": "FFF3CD", "❌ Remaining": "F8D7DA", "⚠️ Failed": "F5B7B1"}

    used_courses = set()

    # ✅ Count how many slots of each kind exist
    slot_occurrences = Counter()
    for y, sems in study_plan_data.items():
        for s, courses in sems.items():
            for c in courses:
                code = canonical_slot(norm(c.get("course_code", "")))
                if is_slot(code): slot_occurrences[code] += 1

    def slot_status(code):
        cslot = canonical_slot(code)

        # Non-slot courses
        if not is_slot(cslot):
            if cslot in completed:   return "✅ Completed", ""
            if cslot in inprogress:  return "🕓 In Progress", ""
            if cslot in failed:      return "⚠️ Failed", ""
            return "❌ Remaining", ""

        # slot courses with explicit mapping
        pools = [(slot_mapping_completed, "✅ Completed"), (slot_mapping_inprog, "🕓 In Progress")]
        for pool, label in pools:
            if pool.get(cslot):
                real = pool[cslot].pop(0)
                used_courses.add(real)
                return label, f" → {real}"

        # fallback to equivalence
        for eq in EQUIVALENT_MAP.get(cslot, []):
            if eq in completed and eq not in used_courses:
                used_courses.add(eq)
                return "✅ Completed", f" → {eq}"
            if eq in inprogress and eq not in used_courses:
                used_courses.add(eq)
                return "🕓 In Progress", f" → {eq}"

        return "❌ Remaining", ""

    # ✅ Sheet 2 (Plan)
    row = 1
    for year, sems in study_plan_data.items():
        for sem, courses in sems.items():
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            h = ws2.cell(row=row, column=1, value=f"{year}: {sem}")
            h.font = Font(bold=True, size=14); h.alignment = Alignment(horizontal="center")
            h.fill = PatternFill("solid", fgColor="D9EAD3")
            row += 1

            headers = ["Course Code", "Course Title", "Credits", "Prerequisite", "Co-Requisite", "Status"]
            for i, col in enumerate(headers, 1):
                c = ws2.cell(row=row, column=i, value=col)
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="E2EFDA"); c.border = border
            row += 1

            total_credits = 0
            for c in courses:
                code = canonical_slot(norm(c.get("course_code", "")))
                if code in removed_plan: continue

                status, note = slot_status(code)
                title = c.get("title", "") + (" " + note if note else "")
                credit = c.get("credits", "")

                try: total_credits += float(credit or 0)
                except: pass

                row_data = [code, title, credit, c.get("prerequisite","-"), c.get("co_requisite","-"), status]
                for i, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=row, column=i, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    if i == 6:
                        cell.fill = PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))
                row += 1

            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws2.cell(row=row, column=1, value="Total Credits:").font = Font(bold=True)
            ws2.cell(row=row, column=6, value=total_credits).font = Font(bold=True)
            row += 2

    # ✅ Legend
    legend = ["Legend:", "✅ Completed", "🕓 In Progress", "❌ Remaining", "⚠️ Failed"]
    ws2.append(legend)
    lr = ws2.max_row
    for i, lbl in enumerate(legend[1:], 2):
        c = ws2.cell(row=lr, column=i, value=lbl)
        c.fill = PatternFill("solid", fgColor=colors[lbl])
        c.border = border
        c.alignment = Alignment(horizontal="center")

    for col in range(1, 8):
        ws2.column_dimensions[chr(64+col)].width = 18

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"💾 Excel saved → {output_path}")
    return output_path

# ---------- Program picking + structuring helpers ----------
def _levels_to_year_sem(levels_dict: dict) -> dict:
    """
    Map numeric levels (1..8) to 'Year / Semester' buckets.
    Handles both string and integer level keys from JSON.
    """
    mapping = {
        "1": ("First Year", "First Semester"),
        "2": ("First Year", "Second Semester"),
        "3": ("Second Year", "First Semester"),
        "4": ("Second Year", "Second Semester"),
        "5": ("Third Year", "First Semester"),
        "6": ("Third Year", "Second Semester"),
        "7": ("Fourth Year", "First Semester"),
        "8": ("Fourth Year", "Second Semester"),
    }

    # ✅ Convert JSON level keys to strings (handles int keys too)
    levels_dict = {str(k): v for k, v in levels_dict.items()}

    structured = {}
    for lvl_key, (year, sem) in mapping.items():
        if year not in structured:
            structured[year] = {}
        structured[year][sem] = levels_dict.get(lvl_key, [])

    return structured

def _pick_program(programs: dict, major_name: str, df_transcript) -> tuple[str, dict]:
    """
    Resolve the correct study plan program:
    1) Exact match (case-insensitive)
    2) Substring match (case-insensitive)
    3) Most overlapping course codes from transcript
    ✅ Handles JSON keys safely
    ✅ Ensures fallback never returns None
    """
    def norm_title(s: str) -> str:
        return " ".join(str(s or "").lower().strip().split())

    # ✅ Normalize incoming major
    major_norm = norm_title(major_name)

    # 1️⃣ Exact match
    if major_norm:
        for k in programs.keys():
            if norm_title(k) == major_norm:
                return k, programs[k]

    # 2️⃣ Substring match
    if major_norm:
        for k in programs.keys():
            if major_norm in norm_title(k):
                return k, programs[k]

    # 3️⃣ Course overlap heuristic
    try:
        student_codes = {
            "".join(str(c).upper().split())
            for c in df_transcript["Course Code"].dropna().tolist()
        }
    except Exception:
        student_codes = set()

    best_key, best_hit = None, -1

    for k, plan in programs.items():
        try:
            plan_codes = {
                "".join(c["course_code"].upper().split())
                for lvl in plan.get("levels", {}).values()
                for c in lvl
            }
        except Exception:
            continue

        hit = len(student_codes & plan_codes)
        if hit > best_hit:
            best_hit, best_key = hit, k

    # ✅ Safety fallback — always return something
    if best_key:
        return best_key, programs[best_key]

    # Last resort: return first program in JSON
    default_key = next(iter(programs.keys()))
    return default_key, programs[default_key]

# ----------------------------------------------------------
# LLM Generator
def generate_llm_response(prompt: str) -> str:
    """Gemini-backed text generation wrapper."""
    from shared_tools import generate_llm_response as _shared_generate_llm_response

    return _shared_generate_llm_response(prompt)
# ==============================================================
# 🎓 Transcript Upload + GPA Conditional Alerts
# ==============================================================
from flask import jsonify, request
import uuid
import pandas as pd
from agents_runtime import _save_state
from pdf_extractor import extract_text, clean_text, parse_courses_with_multiline_fix, extract_gpa_summary_v3
from shared_tools import compare_transcript_with_plan

@app.route("/upload", methods=["POST"])
def upload_transcript():
    """Handle transcript upload, GPA extraction, and dynamic alerts."""
    try:
        if "file" not in request.files:
            return jsonify({"response": "❌ Please upload a transcript PDF."}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"response": "❌ No selected file."}), 400

        # --- Save uploaded file ---
        session_id = str(uuid.uuid4())
        upload_path = os.path.join(UPLOAD_FOLDER, f"{session_id}.pdf")
        file.save(upload_path)
        print(f"📄 Uploaded transcript saved to {upload_path}")

        # --- Extract data ---
        text = clean_text(extract_text(upload_path))
        df_courses = parse_courses_with_multiline_fix(text)

        # --- Extract GPA Summary ---
        prep_text, ug_text = "", text
        df_gpa = extract_gpa_summary_v3(ug_text)
        summary_df = pd.DataFrame(df_gpa)

        # --- Compare with study plan ---
        major = recognize_major_by_overlap(df_courses, STUDY_PLANS)
        summary = compare_transcript_with_plan(
            major, df_courses, {"programs": STUDY_PLANS}, extracted_summary_df=summary_df
        )

        # --- GPA condition checks ---
        gpa_final = summary.get("gpa_final", 0)
        alert_type = None
        message = ""
        standing = ""

        if gpa_final < 2.0:
            alert_type = "warning"
            standing = "⚠️ At Risk (Below 2.0 GPA)"
            message = (
                f"⚠️ Warning: Your GPA is {gpa_final:.2f}, which is below 2.0. "
                "Please contact your academic advisor for guidance."
            )
        elif gpa_final < 3.0:
            alert_type = "encourage"
            standing = "💡 Keep Going (GPA between 2.0 and 3.0)"
            message = (
                f"💡 Good job! Your GPA is {gpa_final:.2f}. Keep improving and aim for 3.0+!"
            )
        else:
            alert_type = "excellent"
            standing = "🌟 Excellent (GPA 3.0+)"
            message = (
                f"🌟 Outstanding! Your GPA is {gpa_final:.2f}. Keep up the amazing work!"
            )

        if gpa_final < ACADEMIC_POLICY["graduation_requirements"]["minimum_gpa"]:
            alert = "⚠️ GPA below graduation threshold"
        elif gpa_final >= ACADEMIC_POLICY["honors"]["first_class"]["min"]:
            alert = "🎓 Eligible for First-Class Honors"


        summary["academic_standing"] = standing

        # --- Save student session ---
        student = {
            "student_name": "Unknown",
            "major": major,
            "gpa_final": gpa_final,
            "progress_percent": summary.get("progress_percent", 0),
            "completed_courses": summary.get("completed_courses", []),
            "remaining_courses": summary.get("remaining_courses", []),
            "academic_standing": standing,
        }

        for k, v in student.items():
            _save_state(session_id, k, v)

        # --- Final response ---
        response_msg = f"✅ Transcript processed.\n🎓 Major: {major}\n📊 GPA: {gpa_final:.2f}\n{message}"

        return jsonify({
        "session_id": session_id,
        "response": response_msg,
        "gpa": gpa_final,
        "major": major,
        "progress_percent": summary.get("progress_percent", 0),
        "alert_type": alert_type,
        "standing": standing,
        "alert": alert  # ✅ Add this line
    })


    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"response": f"❌ Error processing file: {e}"}), 500

# ----------------------------------------------------------
# ROUTES
# ----------------------------------------------------------

@app.route("/chat", methods=["POST"])
def chat():
    from agents_runtime import load_student_state
    
    data = request.json or {}
    message = (data.get("message") or "").strip()

    if not message:
        return jsonify({"response": "Please enter a question."})

    # Get session id
    session_id = request.cookies.get("session_id")
    st = load_student_state(session_id) if session_id else None

    if not st:
        return jsonify({"response":"📎 Please upload your transcript first."})

    # Clean placeholder course slots
    def clean(courses):
        ignore = {"AIXXX","C3SXXX","GHALXXX","GIASXXX","GSOSXXX"}
        return [c for c in (courses or []) if c not in ignore]

    facts = {
        "name": st.name or "Student",
        "major": st.major or "Not detected",
        "gpa_prep": st.gpa_prep,
        "gpa_ug": st.gpa_undergrad,
        "progress": st.progress_percent,
        "completed": clean(st.completed_courses),
        "remaining": clean(st.remaining_courses),
        "in_progress": clean(st.in_progress_courses)
    }

    # Arabic or English
    is_ar = bool(re.search(r"[\u0600-\u06FF]", message))
    lang = "Arabic" if is_ar else "English"

    # Prompt (simple — NO ReAct)
    prompt = f"""
You are UniMate, the official academic advisor of Prince Muqrin University.

Student Data:
{json.dumps(facts, ensure_ascii=False, indent=2)}

User question:
"{message}"

Rules:
- Respond only in {lang}
- Use ONLY the above student data
- Do NOT invent course names
- If asked about remaining courses, list them
- If asked GPA, use stored values
- If question not academic → give supportive response
- Be concise, friendly, professional

Answer:
"""

    reply = generate_llm_response(prompt).strip()
    if "gpa" in message.lower():
        alert_msg = get_alert_message(facts)
        if alert_msg:
            reply += f"\n\n{alert_msg}"
    return jsonify({"response": reply})
# advisor_alerts.py
# advisor_alerts.py
def get_alert_message(facts):
    """
    Returns a short alert message about GPA or academic progress
    based on student's record.
    """
    # Safely cast GPA and progress to float
    try:
        gpa = float(facts.get("gpa_ug", 0) or 0)
    except (ValueError, TypeError):
        gpa = 0.0

    try:
        progress = float(facts.get("progress", 0) or 0)
    except (ValueError, TypeError):
        progress = 0.0

    # GPA-based alerts
    if gpa < 2.0:
        return f"⚠️ Your GPA is {gpa:.2f}, which is below 2.0. Please meet your academic advisor for a recovery plan."
    elif gpa < 2.5:
        return f"💡 Your GPA is {gpa:.2f}. Consider retaking low-grade courses to improve your standing."
    elif gpa >= 3.75:
        return f"🌟 Excellent work! Your GPA is {gpa:.2f}. You are eligible for the Honor List this semester."

    # Progress-based alerts
    if progress >= 85 and progress < 100:
        return "🎯 You are close to graduation! Make sure you’ve completed all required electives and co-op/internship."
    elif progress == 100:
        return "🎓 Congratulations! You have completed all your degree requirements."

    return None


# ----------------------------------------------------------
# DOWNLOAD EXCEL REPORT
# ----------------------------------------------------------
@app.route("/download-report/<filename>")
def download_report(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)

    if not os.path.exists(file_path):
        return "File not found", 404

    @after_this_request
    def cleanup(response):
        # ❗ Leave file in uploads (remove later if needed)
        return response

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ----------------------------------------------------------
# Serve uploaded files if needed
# ----------------------------------------------------------
@app.route("/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


# ----------------------------------------------------------
# AGENTS UPLOAD ENDPOINT
# ----------------------------------------------------------
def get_agent_handlers():
    from agents_runtime import ui_agent_handle_upload
    return ui_agent_handle_upload


@app.route("/agents/upload", methods=["POST"])
def agents_upload():
    from agents_runtime import StudentState, save_student_state
    ui_agent_handle_upload = get_agent_handlers()

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)

    # 🔁 Agent pipeline upload
    session_id, student, plan, ui_note = ui_agent_handle_upload(
        path,
        user_id="local_user",
        reasoning_mode=request.args.get("mode", "react+reflexion"),
        hitl=False
    )

    # ------------------------------------------------------
    # ✅ Create & Save Structured StudentState
    # ------------------------------------------------------
    student_state = StudentState(
        session_id=session_id,
        name=getattr(student, "name", None),
        major=getattr(plan, "major", None),
        gpa_prep=plan.summary.get("gpa_prep"),
        gpa_undergrad=plan.summary.get("gpa_undergrad") or plan.summary.get("gpa_final"),
        credits_prep=plan.summary.get("credit_prep"),
        credits_undergrad=plan.summary.get("credit_ug"),
        credits_in_progress=plan.summary.get("credit_inprogress"),
        progress_percent=plan.summary.get("progress_percent"),
        completed_courses=plan.summary.get("completed_courses", []),
        in_progress_courses=plan.summary.get("in_progress_courses", []),
        remaining_courses=plan.summary.get("remaining_courses", []),
    )

    # Save to DB
    save_student_state(student_state)
    # ------------------------------------------------------
    # 🎓 Smart GPA + Progress Alert System (Policy-Driven)
    # ------------------------------------------------------
    try:
        gpa_final = float(plan.summary.get("gpa_undergrad") or plan.summary.get("gpa_final") or 0)
    except Exception:
        gpa_final = 0.0

    try:
        progress = float(plan.summary.get("progress_percent") or 0)
    except Exception:
        progress = 0.0

    alert_type = None
    alert_message = None
    standing = ""

    # 🔹 Load GPA rules dynamically from academic_policy.json
    rules = ACADEMIC_POLICY.get("gpa_rules", {})
    honors = ACADEMIC_POLICY.get("honors", {})
    grad_req = ACADEMIC_POLICY.get("graduation_requirements", {})

    # --- GPA-based classification (policy driven) ---
    if "fail" in rules and gpa_final <= rules["fail"].get("max", 1.99):
        alert_type = "warning"
        standing = "⚠️ Academic Warning"
        alert_message = f"⚠️ GPA {gpa_final:.2f}: Below {rules['pass']['min']:.2f}. You are at risk of probation."
    elif "pass" in rules and rules["pass"]["min"] <= gpa_final <= rules["pass"]["max"]:
        alert_type = "encourage"
        standing = "💡 Pass Standing"
        alert_message = f"💡 GPA {gpa_final:.2f}: Minimum passing range. Focus on steady improvement."
    elif "good" in rules and rules["good"]["min"] <= gpa_final <= rules["good"]["max"]:
        alert_type = "encourage"
        standing = "👍 Good Standing"
        alert_message = f"👍 GPA {gpa_final:.2f}: Good academic standing. Keep pushing for higher distinction."
    elif "very_good" in rules and rules["very_good"]["min"] <= gpa_final <= rules["very_good"]["max"]:
        alert_type = "encourage"
        standing = "🌟 Very Good Standing"
        alert_message = f"🌟 GPA {gpa_final:.2f}: Excellent consistency. You're close to First-Class Honors!"
    elif "excellent" in rules and rules["excellent"]["min"] <= gpa_final <= rules["excellent"]["max"]:
        alert_type = "excellent"
        standing = "🎓 First-Class Honor"
        alert_message = f"🎓 GPA {gpa_final:.2f}: Outstanding performance. You qualify for First-Class Honors!"

    # --- Graduation & honors checks from policy file ---
    if gpa_final < grad_req.get("minimum_gpa", 2.0):
        alert_message += " ⚠️ GPA below graduation threshold."
    elif gpa_final >= honors["first_class"]["min"]:
        alert_message += " 🎓 Eligible for First-Class Honors."
    elif gpa_final >= honors["second_class"]["min"]:
        alert_message += " 🏅 Eligible for Second-Class Honors."

    # --- Add progress-based note ---
    if progress >= 85 and progress < 100:
        alert_message += " 🎯 You are close to graduation — ensure electives & co-op are completed."
    elif progress == 100:
        alert_message += " 🎓 Congratulations! You have completed all degree requirements."

    # --- Optional contextual encouragement ---
    if not alert_message:
        alert_type = "info"
        alert_message = f"💡 GPA {gpa_final:.2f}: Keep striving for excellence!"

    # 🔍 Debug output
    print(f"[DEBUG] GPA={gpa_final}, Progress={progress}, Standing={standing}, AlertType={alert_type}")


    # ------------------------------------------------------
    # Build file download link
    # ------------------------------------------------------
    from os.path import basename
    excel_url = f"/download-report/{basename(plan.excel_path)}" if plan.excel_path else None

    # ------------------------------------------------------
    # ✅ Final response (includes alerts)
    # ------------------------------------------------------
    resp = jsonify({
        "session_id": session_id,
        "major": plan.major,
        "excel_path": excel_url,
        "ui_summary": ui_note,
        "progress_percent": plan.summary.get("progress_percent"),
        "remaining_courses": plan.summary.get("remaining_courses", []),
        "gpa": gpa_final,
        "alert_type": alert_type,
        "alert": alert_message
    })

    # ------------------------------------------------------
    # Set session cookie for chat
    # ------------------------------------------------------
    resp.set_cookie(
        "session_id",
        session_id,
        httponly=True,
        samesite="Lax",
        secure=False  # ✅ True only in production HTTPS
    )

    return resp


@app.route("/debug-gpa", methods=["POST"])
def debug_gpa():
    """
    Debug GPA detection directly from a PDF upload without processing the entire transcript.
    Returns all detected 'Cum GPA' values for inspection.
    """
    import re
    from pdf_extractor import extract_text, clean_text, split_academic_blocks

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)

    try:
        # Extract full text
        text = clean_text(extract_text(path))
        prep_text, ug_text = split_academic_blocks(text)

        # Find all GPA values
        matches = re.findall(r"Cum GPA\s*([\d.]+)\s*Cumulative Total", ug_text)
        raw_matches = re.findall(r"(Cum GPA[^\n]*)", ug_text)

        response = {
            "total_matches_found": len(matches),
            "all_matches": matches,
            "raw_lines": raw_matches[-10:],  # show last few lines around GPA mentions
            "last_detected_gpa": float(matches[-1]) if matches else None
        }

        print("📄 GPA DEBUG INFO")
        print("=" * 60)
        print(f"Found {len(matches)} matches: {matches}")
        if matches:
            print(f"✅ Final detected GPA: {matches[-1]}")
        else:
            print("⚠️ No GPA pattern matched.")
        print("=" * 60)

        return jsonify(response)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    import webbrowser
    try:
        debug_log("🚀 Starting Flask server from main block...")
        webbrowser.open("http://127.0.0.1:5000")
        app.run(debug=False)
    except Exception:
        debug_log("💥 Startup error:\n" + traceback.format_exc())
        input("Press Enter to close...")
