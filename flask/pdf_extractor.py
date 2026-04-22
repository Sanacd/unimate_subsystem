import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# =========================================================
# 1️⃣ TEXT UTILITIES
# =========================================================
def extract_text(pdf_path: str) -> str:
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += "\n" + (page.extract_text() or "")
    return text

def clean_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()

# =========================================================
# 2️⃣ SPLIT BLOCKS: Preparatory vs Undergraduate
# =========================================================
def split_academic_blocks(text: str):
    """Split transcript text into preparatory and undergraduate sections."""
    def rx(words):
        return r"\s*".join(map(re.escape, words))

    PREP_BEGIN = rf"{rx(['Beginning','of','Preparatory','Year','Record'])}s?"
    PREP_END   = rf"{rx(['End','of','Preparatory','Year','Record'])}s?"
    UG_BEGIN   = rf"{rx(['Beginning','of','Undergraduate','Record'])}s?"
    UG_END     = rf"{rx(['End','of','Undergraduate','Record'])}s?"

    flags = re.I | re.S
    prep_text, ug_text = "", ""

    prep_match = re.search(f"{PREP_BEGIN}(.*?){PREP_END}", text, flags)
    if prep_match:
        prep_text = prep_match.group(1)

    ug_match = re.search(f"{UG_BEGIN}(.*?){UG_END}", text, flags)
    if ug_match:
        ug_text = ug_match.group(1)
    else:
        ug_match2 = re.search(f"{UG_BEGIN}(.*)$", text, flags)
        if ug_match2:
            ug_text = ug_match2.group(1)

    if not prep_text and not ug_text:
        ug_text = text

    return prep_text, ug_text


# =========================================================
# 3️⃣ STUDENT INFO
# =========================================================
def parse_student_info_v3(text: str):
    info = {}
    patterns = {
        "Name": r"Name:\s*([A-Za-z ,.'-]+)",
        "Student ID": r"Student\s*ID:\s*(\d+)",
        "Date of Birth": r"Date\s*of\s*Birth:\s*([\d/]+)",
        "Date Issued": r"Date\s*Issued:\s*([\d/]+)",
        "Nationality": r"Nationality:\s*([A-Za-z ]+)",
        "Cumulative GPA": r"Cumulative\s*GPA:\s*([\d.]+)",
    }

    def _clean_name(raw):
        if not raw:
            return raw

        # Remove transcript header junk
        junk_words = [
            "Admission Date",
            "University Of Prince Mugrin",
            "University of Prince Mugrin",
            "University",
        ]
        for jw in junk_words:
            if jw in raw:
                raw = raw.split(jw)[0]

        # Remove too many spaces/commas
        raw = raw.strip().rstrip(",")
        return raw

    # Extract based on patterns
    for key, pat in patterns.items():
        m = re.search(pat, text, re.I)
        if m:
            info[key] = m.group(1).strip()

    # Prep / Undergraduate college detection fallback
    college_prep = re.search(r"Preparatory\s+Year", text, re.I)
    college_ug = re.search(r"College\s+of\s+[A-Za-z &]+", text, re.I)

    info["Preparatory College"] = "Preparatory Year" if college_prep else None

    if college_ug:
        info["Undergraduate College"] = (
            college_ug.group(0).replace(" CS", "").strip()
        )
    else:
        info["Undergraduate College"] = "College of Computer and Cyber Sciences"

    # ✅ Clean Name BEFORE title-case
    raw_name = info.get("Name", "")
    clean_name = _clean_name(raw_name)
    info["Name"] = clean_name.title() if clean_name else ""

    # ✅ Format Nationality softly
    info["Nationality"] = info.get("Nationality", "").title() if info.get("Nationality") else ""

    return info


# =========================================================
# 4️⃣ COURSE PARSER
# =========================================================
def parse_courses_with_multiline_fix(text: str) -> pd.DataFrame:
    t = re.sub(r"(\w)\n(\w)", r"\1 \2", text)
    t = re.sub(r"(\n\s*){2,}", "\n", t)

    lines = t.splitlines()
    fixed_lines, buffer_line = [], ""
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if not re.match(r"^[A-Z]{2,4}\s*\d{3}", stripped) and not re.search(r"\d+\.\d+", stripped):
            buffer_line += " " + stripped
        else:
            if buffer_line and fixed_lines:
                fixed_lines[-1] += " " + buffer_line.strip()
                buffer_line = ""
            fixed_lines.append(stripped)
    if buffer_line and fixed_lines:
        fixed_lines[-1] += " " + buffer_line.strip()

    t_fixed = "\n".join(fixed_lines)
    pat = re.compile(
        r"([A-Z]{2,4}\s*\d{3})\s+([A-Za-z0-9 ,()’'–\-/]+?)\s+(\d+\.\d+)\s+(?:([A-Z\+]+|Waived|0\.00|0\.0|W))?\s*(\d+\.\d+)?",
        re.M,
    )
    rows = []
    for m in pat.finditer(t_fixed):
        rows.append({
            "Course Code": m.group(1).strip(),
            "Course Title": re.sub(r"\s+", " ", m.group(2)).strip().title(),
            "Credit Hours": float(m.group(3)),
            "Grade": (m.group(4) or "").strip(),
            "Points": float(m.group(5)) if m.group(5) else 0.0,
        })
    return pd.DataFrame(rows)


# =========================================================
# 5️⃣ TITLE NORMALIZER
# =========================================================
def smart_fix_titles(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize partial or split course titles into full consistent names."""
    fixes = {
        "Data Structures And": "Data Structures and Algorithms",
        "Web Application": "Web Application Development",
        "Introduction To Discrete": "Introduction to Discrete Structures",
        "Operating": "Fundamentals of Operating Systems",
        "Principles Of Physics": "Principles of Physics I",
    }
    for partial, full in fixes.items():
        df.loc[df["Course Title"].str.startswith(partial, na=False), "Course Title"] = full
    return df


# =========================================================
# 6️⃣ CATEGORY SPLIT
# =========================================================
def split_by_category_v4(df: pd.DataFrame):
    prep_codes = {
        "ENGL 001", "ENGL 002", "ENGL 003", "ENGL 004", "ENGL 005",
        "MATH 001", "MATH 002", "PCS 001", "PCD 001"
    }

    df["Category"] = "Undergraduate"
    df.loc[df["Course Code"].isin(prep_codes), "Category"] = "Preparatory"
    df.loc[df["Grade"].astype(str).str.fullmatch(r"(W|Waived)", case=False, na=False), "Category"] = "Waived"
    df.loc[
        ((df["Points"] == 0.0) | (df["Points"].isna())) &
        (df["Grade"].astype(str).isin(["", "0", "0.0", "0.00"])) ,
        "Category"
    ] = "In Progress"

    df.loc[df["Course Code"].str.startswith(("C3S", "GHAL","GIAS","GSOS","GDMC")), "Category"] = "Undergraduate"

    return (
        df[df["Category"] == "Preparatory"].reset_index(drop=True),
        df[df["Category"] == "Undergraduate"].reset_index(drop=True),
        df[df["Category"] == "In Progress"].reset_index(drop=True),
        df[df["Category"] == "Waived"].reset_index(drop=True)
    )


# =========================================================
# 7️⃣ UNIVERSAL FINAL UNDERGRADUATE GPA DETECTION (DEBUG-ENHANCED)
# =========================================================
def extract_final_ug_gpa(text: str, debug: bool = True) -> float | None:
    """Detect final UG GPA by matching the highest credit total rather than last match."""
    _, ug_text = split_academic_blocks(text)
    section = ug_text or text

    pattern = re.compile(
        r"Cum\s*GPA\s*([\d.]+)\s*Cumulative\s*Total\s*([\d.]+)\s*([\d.]+)",
        re.I
    )
    matches = list(pattern.finditer(section))

    if not matches:
        if debug:
            print("⚠️ No GPA/Cumulative Total pattern found.")
        return None

    gpa_records = []
    for i, m in enumerate(matches, 1):
        gpa_val = float(m.group(1))
        credits = float(m.group(2))
        points = float(m.group(3))
        gpa_records.append((credits, gpa_val, points))
        if debug:
            context = section[max(0, m.start() - 60):min(len(section), m.end() + 60)]
            print(f"[{i}] GPA={gpa_val}, Credits={credits}, Points={points}")
            print(f"   🧠 Context: ...{context.replace(chr(10), ' ')}...")

    # Pick the GPA with the highest total credits — most likely the final overall
    final_record = max(gpa_records, key=lambda x: x[0])
    final_gpa = final_record[1]

    if debug:
        print(f"🎯 Selected Final GPA: {final_gpa} (based on highest total credits {final_record[0]})")

    return final_gpa



# =========================================================
# 8️⃣ GPA EXTRACTION (Document Order)
# =========================================================
def extract_gpa_summary_v3(text: str) -> pd.DataFrame:
    text = re.sub(r"(\n\s*){2,}", "\n", text)
    pattern = re.compile(
        r"Acad\s*Year\s+(\d{4}\s*-\s*\d{4})\s+(Fall|Spring|Summer).*?Sem\s*GPA\s*([\d.]+).*?Cum\s*GPA\s*([\d.]+)",
        re.S | re.I,
    )
    records = []
    for m in pattern.finditer(text):
        records.append({
            "_pos": m.start(),
            "Academic Year": m.group(1).strip(),
            "Semester": m.group(2).strip().title(),
            "Semester GPA": float(m.group(3)),
            "Cumulative GPA": float(m.group(4)),
        })
    if not records:
        return pd.DataFrame(columns=["Academic Year", "Semester", "Semester GPA", "Cumulative GPA"])
    return (
        pd.DataFrame(records)
        .sort_values("_pos")
        .drop(columns="_pos")
        .drop_duplicates(subset=["Academic Year", "Semester"], keep="first")
        .reset_index(drop=True)
    )

# =========================================================
# 9️⃣ SUMMARY CREATION (DEBUG-ENHANCED)
# =========================================================
def create_summary(student_info, df_all, df_prep, df_ug, df_inprog, df_waived,
                   df_gpa_prep, df_gpa_ug, full_text=None, debug: bool = True):
    total_credits = df_all["Credit Hours"].sum()

    # Preparatory GPA
    prep_gpa = None
    if isinstance(student_info, dict):
        try:
            prep_gpa = float(student_info.get("Cumulative GPA", "") or "nan")
        except Exception:
            prep_gpa = None
    if prep_gpa is None and not df_gpa_prep.empty:
        prep_gpa = float(df_gpa_prep["Cumulative GPA"].iloc[-1])

    # Undergraduate GPA
    ug_gpa = None
    if full_text:
        ug_gpa = extract_final_ug_gpa(full_text, debug=debug)
    if ug_gpa is None and not df_gpa_ug.empty:
        ug_gpa = float(df_gpa_ug["Cumulative GPA"].iloc[-1])

    if debug:
        print("\n📊 [DEBUG] GPA SUMMARY:")
        print(f"   Preparatory GPA: {prep_gpa}")
        print(f"   Final UG GPA: {ug_gpa}")
        print(f"   Total Credits: {total_credits}")

    summary = {
        "Total Courses": len(df_all),
        "Total Credit Hours": float(total_credits),
        "Preparatory Credits": float(df_prep["Credit Hours"].sum()),
        "Undergraduate Credits": float(df_ug["Credit Hours"].sum()),
        "In-Progress Credits": float(df_inprog["Credit Hours"].sum()),
        "Waived Credits": float(df_waived["Credit Hours"].sum()),
        "Preparatory GPA": prep_gpa,
        "Final Cumulative GPA": ug_gpa
    }
    return pd.DataFrame(list(summary.items()), columns=["Metric", "Value"])
# =========================================================
# 🔟 EXCEL EXPORT
# =========================================================
def style_excel(output_path):
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    wb.save(output_path)
    print(" Styling applied successfully.")


def save_to_excel(student_info, df_all, df_prep, df_ug, df_inprog, df_waived,
                  df_gpa_prep, df_gpa_ug, df_summary, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame([student_info]).T.rename(columns={0: "Value"}).to_excel(writer, sheet_name="Student_Info")
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_all.to_excel(writer, sheet_name="All_Courses", index=False)
        df_prep.to_excel(writer, sheet_name="Preparatory_Courses", index=False)
        df_ug.to_excel(writer, sheet_name="Undergraduate_Courses", index=False)
        df_inprog.to_excel(writer, sheet_name="In_Progress", index=False)
        df_waived.to_excel(writer, sheet_name="Waived_Courses", index=False)
        if not df_gpa_prep.empty:
            df_gpa_prep.to_excel(writer, sheet_name="GPA_Prep", index=False)
        if not df_gpa_ug.empty:
            df_gpa_ug.to_excel(writer, sheet_name="GPA_Undergrad", index=False)
    style_excel(output_path)
    print(f"✅ Excel file created successfully → {output_path}")


def _extract_text_from_supported_file(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return clean_text(extract_text(file_path))
    with open(file_path, "r", encoding="utf-8") as fh:
        return clean_text(fh.read())


def _extract_term_markers(text: str) -> list[tuple[int, str]]:
    markers = []
    patterns = [
        re.compile(r"Acad\s*Year\s+(\d{4}\s*-\s*\d{4})\s+(Fall|Spring|Summer)", re.I),
        re.compile(r"\b(Fall|Spring|Summer)\s+(20\d{2})\b", re.I),
    ]

    for pattern in patterns:
        for match in pattern.finditer(text):
            if "Acad" in pattern.pattern:
                term = f"{match.group(2).title()} {match.group(1).strip()}"
            else:
                term = f"{match.group(1).title()} {match.group(2)}"
            markers.append((match.start(), term))

    return sorted(markers, key=lambda item: item[0])


def _term_for_position(position: int, markers: list[tuple[int, str]]) -> str:
    term = ""
    for marker_pos, marker_term in markers:
        if marker_pos <= position:
            term = marker_term
        else:
            break
    return term


def _classify_transcript_status(grade_raw, points_raw) -> str:
    grade_text = str(grade_raw or "").strip().upper()
    if grade_text in {"IP", "I", "IN PROGRESS", "INPROGRESS"}:
        return "in_progress"
    if grade_text in {"F", "FA", "NF"}:
        return "failed"
    if grade_text in {"W", "WF", "WP", "WITHDRAWN"}:
        return "not_taken"
    try:
        if float(points_raw or 0) > 0:
            return "completed"
    except (TypeError, ValueError):
        pass
    if grade_text and grade_text not in {"0", "0.0", "0.00"}:
        return "completed"
    return "in_progress"


def extract_transcript_data(file_path: str) -> dict:
    text = _extract_text_from_supported_file(file_path)
    student_info = parse_student_info_v3(text)
    term_markers = _extract_term_markers(text)

    course_pattern = re.compile(
        r"([A-Z]{2,6}\s*[-]?\s*\d{3}[A-Z]?)\s+([A-Za-z0-9 ,()â€™'â€“\-/]+?)\s+(\d+(?:\.\d+)?)\s+(?:([A-Z\+]+|Waived|0\.00|0\.0|W|IP|I))?\s*(\d+(?:\.\d+)?)?",
        re.M,
    )

    courses = []
    for match in course_pattern.finditer(text):
        code = re.sub(r"[^A-Z0-9]", "", match.group(1).upper())
        title = re.sub(r"\s+", " ", match.group(2)).strip().title()
        try:
            credits = float(match.group(3))
        except (TypeError, ValueError):
            credits = 0.0
        grade = (match.group(4) or "").strip()
        try:
            points = float(match.group(5)) if match.group(5) else 0.0
        except (TypeError, ValueError):
            points = 0.0

        courses.append(
            {
                "course_code": code,
                "course_name": title,
                "credits": credits,
                "grade": grade,
                "points": points,
                "status": _classify_transcript_status(grade, points),
                "term_taken": _term_for_position(match.start(), term_markers),
                "notes": "",
            }
        )

    if not courses:
        df_courses = parse_courses_with_multiline_fix(text)
        for row in df_courses.to_dict("records"):
            code = re.sub(r"[^A-Z0-9]", "", str(row.get("Course Code") or "").upper())
            grade = str(row.get("Grade") or "").strip()
            points = row.get("Points") or 0.0
            courses.append(
                {
                    "course_code": code,
                    "course_name": str(row.get("Course Title") or "").strip(),
                    "credits": float(row.get("Credit Hours") or 0),
                    "grade": grade,
                    "points": float(points or 0),
                    "status": _classify_transcript_status(grade, points),
                    "term_taken": "",
                    "notes": "",
                }
            )

    gpa_df = extract_gpa_summary_v3(text)
    gpa_final = None
    if not gpa_df.empty:
        try:
            gpa_final = float(gpa_df["Cumulative GPA"].astype(float).iloc[-1])
        except Exception:
            gpa_final = None

    return {
        "student": {
            "student_name": student_info.get("Name") or "",
            "student_id": student_info.get("Student ID") or "",
            "program": student_info.get("Undergraduate College") or "",
            "gpa_final": gpa_final,
        },
        "courses": courses,
        "gpa_table": gpa_df.to_dict("records") if hasattr(gpa_df, "to_dict") else [],
    }


# =========================================================
# 🧩 DEBUG UTILITIES
# =========================================================
def debug_gpa_detection(pdf_path: str):
    """Prints out the text around GPA detections for debugging."""
    print(f"\n🔍 DEBUGGING GPA DETECTION FOR: {pdf_path}\n{'='*60}")
    text = clean_text(extract_text(pdf_path))
    prep_text, ug_text = split_academic_blocks(text)
    section = ug_text or text

    # Detect End Marker
    end_marker = re.findall(r"End\s+of\s+Undergraduate\s+Record[s]?", text, re.I)
    print(f"🪪 Found {len(end_marker)} 'End of Undergraduate Record' markers.")

    # Use universal GPA regex list
    gpa_patterns = [
        r"Cum GPA\s*([\d.]+)\s*Cumulative Total",
        r"Cumulative GPA[:\s]*([\d.]+)",
        r"CUM GPA[:\s]*([\d.]+)",
        r"GPA\s*\(Cumulative\)\s*[:\s]*([\d.]+)",
        r"Cumulative Grade Point Average[:\s]*([\d.]+)",
    ]

    matches = []
    for pat in gpa_patterns:
        for m in re.finditer(pat, section, re.I):
            matches.append(m)
            print(f"🔎 Pattern '{pat}' matched → {m.group(1)}")

    if not matches:
        print("⚠️ No Cumulative GPA values found — likely formatting issue or OCR mismatch.")
    else:
        print(f"✅ Found {len(matches)} GPA matches, last one likely final = {matches[-1].group(1)}")

    # Show tail of UG section
    print("\n📜 Last 600 characters of UG section:")
    print("=" * 60)
    print(section[-600:])
    print("=" * 60)
